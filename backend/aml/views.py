"""
API Views for AML System
"""
import logging
import csv
import io
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.utils import timezone
from django.db.models import Q
from django.http import FileResponse, HttpResponse
from django.shortcuts import render

from .models import Customer, Transaction, Alert, AlertComment, RiskScore, Rule, Report, AuditLog, Device, Merchant
from .serializers import (
    CustomerSerializer, TransactionSerializer, AlertSerializer,
    RiskScoreSerializer, RuleSerializer, ReportSerializer, AuditLogSerializer,
    MonitorTransactionSerializer, ReviewAlertSerializer, GenerateReportSerializer,
    DeviceSerializer, MerchantSerializer, AlertCommentSerializer, AssignAlertSerializer
)
from .services.transaction_monitor import get_transaction_monitor
from .services.alert_generator import get_alert_generator
from .services.report_generator import get_report_generator

logger = logging.getLogger('aml')


def dashboard_view(request):
    """
    Root UI: Didebaan Fraud & Abuse Intelligence dashboard (counts + charts).
    Issues #18, #19: Dashboard view with alerts overview and risk distribution.
    """
    import json
    from django.db.models import Count

    # Summary counts
    customers_total = Customer.objects.count()
    transactions_total = Transaction.objects.count()
    alerts_open = Alert.objects.filter(status='OPEN').count()
    rules_active = Rule.objects.filter(status='ACTIVE').count()

    # Issue #19: Alert severity distribution (for chart)
    severity_qs = (
        Alert.objects.values('severity')
        .annotate(count=Count('id'))
        .order_by('severity')
    )
    severity_labels = []
    severity_data = []
    severity_colors = {'LOW': '#22c55e', 'MEDIUM': '#f59e0b', 'HIGH': '#ef4444', 'CRITICAL': '#7c3aed'}
    for item in severity_qs:
        severity_labels.append(item['severity'])
        severity_data.append(item['count'])

    # Issue #19: Customer risk distribution (for chart)
    risk_qs = (
        Customer.objects.values('current_risk_level')
        .annotate(count=Count('id'))
        .order_by('current_risk_level')
    )
    risk_labels = []
    risk_data = []
    for item in risk_qs:
        risk_labels.append(item['current_risk_level'])
        risk_data.append(item['count'])

    # Issue #19: Recent 7-day alert trend
    from datetime import timedelta
    from django.db.models.functions import TruncDate
    trend_start = timezone.now() - timedelta(days=7)
    trend_qs = (
        Alert.objects.filter(created_at__gte=trend_start)
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(count=Count('id'))
        .order_by('day')
    )
    trend_labels = [item['day'].strftime('%m/%d') for item in trend_qs]
    trend_data = [item['count'] for item in trend_qs]

    context = {
        'customers': customers_total,
        'transactions': transactions_total,
        'alerts_open': alerts_open,
        'rules': rules_active,
        # Chart data (JSON-safe)
        'severity_labels_json': json.dumps(severity_labels),
        'severity_data_json': json.dumps(severity_data),
        'severity_colors_json': json.dumps([severity_colors.get(l, '#64748b') for l in severity_labels]),
        'risk_labels_json': json.dumps(risk_labels),
        'risk_data_json': json.dumps(risk_data),
        'trend_labels_json': json.dumps(trend_labels),
        'trend_data_json': json.dumps(trend_data),
    }
    return render(request, 'aml/dashboard.html', context)


class CustomerViewSet(viewsets.ModelViewSet):
    """ViewSet for Customer model"""
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'customer_id'
    filterset_fields = ['current_risk_level', 'customer_type', 'is_active', 'country']
    search_fields = ['customer_id', 'first_name', 'last_name', 'email', 'national_id']
    ordering_fields = ['created_at', 'registration_date', 'risk_score', 'current_risk_level']
    ordering = ['-created_at']

    def get_queryset(self):
        queryset = Customer.objects.all()
        
        # Filter by risk level
        risk_level = self.request.query_params.get('risk_level', None)
        if risk_level:
            queryset = queryset.filter(current_risk_level=risk_level)
        
        # Filter by customer type
        customer_type = self.request.query_params.get('customer_type', None)
        if customer_type:
            queryset = queryset.filter(customer_type=customer_type)
        
        # Search by name or email
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(customer_id__icontains=search)
            )
        
        return queryset
    
    @action(detail=True, methods=['get'])
    def risk_scores(self, request, customer_id=None):
        """Get risk scores for a customer"""
        customer = self.get_object()
        risk_scores = RiskScore.objects.filter(customer=customer).order_by('-calculated_at')
        serializer = RiskScoreSerializer(risk_scores, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def alerts(self, request, customer_id=None):
        """Get alerts for a customer"""
        customer = self.get_object()
        alerts = Alert.objects.filter(customer=customer).order_by('-created_at')
        serializer = AlertSerializer(alerts, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def transactions(self, request, customer_id=None):
        """Get transactions for a customer"""
        customer = self.get_object()
        transactions = Transaction.objects.filter(customer=customer).order_by('-transaction_date')
        serializer = TransactionSerializer(transactions, many=True)
        return Response(serializer.data)


class TransactionViewSet(viewsets.ModelViewSet):
    """ViewSet for Transaction model"""
    queryset = Transaction.objects.all()
    serializer_class = TransactionSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'transaction_id'
    filterset_fields = ['transaction_type', 'status', 'is_suspicious', 'currency']
    search_fields = ['transaction_id', 'sender_account', 'receiver_account', 'receiver_name']
    ordering_fields = ['transaction_date', 'amount', 'created_at']
    ordering = ['-transaction_date']

    def get_queryset(self):
        queryset = Transaction.objects.all()
        
        # Filter by suspicious
        is_suspicious = self.request.query_params.get('is_suspicious', None)
        if is_suspicious is not None:
            queryset = queryset.filter(is_suspicious=is_suspicious.lower() == 'true')
        
        # Filter by transaction type
        transaction_type = self.request.query_params.get('transaction_type', None)
        if transaction_type:
            queryset = queryset.filter(transaction_type=transaction_type)
        
        # Filter by date range
        date_from = self.request.query_params.get('date_from', None)
        date_to = self.request.query_params.get('date_to', None)
        if date_from:
            queryset = queryset.filter(transaction_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(transaction_date__lte=date_to)
        
        return queryset
    
    @action(detail=False, methods=['post'], url_path='monitor')
    def monitor_transaction(self, request):
        """Monitor a transaction and generate alerts if needed"""
        serializer = MonitorTransactionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        transaction_id = serializer.validated_data['transaction_id']
        
        try:
            transaction = Transaction.objects.get(transaction_id=transaction_id)
        except Transaction.DoesNotExist:
            return Response(
                {'error': f'Transaction {transaction_id} not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Monitor transaction
        monitor = get_transaction_monitor()
        result = monitor.monitor_transaction(transaction)
        
        # Generate alert if needed
        alert = None
        if result['should_alert']:
            alert_generator = get_alert_generator()
            
            # Get triggered rules
            from .rules.aml_rules import get_rule_engine
            rule_engine = get_rule_engine()
            triggered_rules, reasons, _ = rule_engine.evaluate_transaction(transaction)
            
            alert = alert_generator.generate_alert(
                transaction=transaction,
                triggered_rules=triggered_rules,
                risk_score=result['risk_score'],
                severity=result['alert_severity'],
                reasons=reasons
            )
        
        response_data = {
            'monitoring_result': result,
            'alert': AlertSerializer(alert).data if alert else None
        }
        
        return Response(response_data, status=status.HTTP_200_OK)


class AlertViewSet(viewsets.ModelViewSet):
    """ViewSet for Alert model"""
    queryset = Alert.objects.all()
    serializer_class = AlertSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'alert_id'
    filterset_fields = ['status', 'severity']
    search_fields = ['alert_id', 'title', 'description']
    ordering_fields = ['created_at', 'risk_score', 'severity']
    ordering = ['-created_at']

    def get_queryset(self):
        queryset = Alert.objects.all()
        
        # Filter by status
        alert_status = self.request.query_params.get('status', None)
        if alert_status:
            queryset = queryset.filter(status=alert_status)
        
        # Filter by severity
        severity = self.request.query_params.get('severity', None)
        if severity:
            queryset = queryset.filter(severity=severity)
        
        return queryset
    
    @action(detail=True, methods=['post'], url_path='review')
    def review_alert(self, request, alert_id=None):
        """Review an alert"""
        alert = self.get_object()
        serializer = ReviewAlertSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        alert_generator = get_alert_generator()
        reviewer = request.user.username if hasattr(request.user, 'username') else 'system'
        
        status_value = serializer.validated_data['status']
        notes = serializer.validated_data['notes']
        
        if status_value == 'ESCALATED':
            alert = alert_generator.escalate_alert(alert, reviewer, notes)
        elif status_value == 'FALSE_POSITIVE':
            alert = alert_generator.mark_false_positive(alert, reviewer, notes)
        else:
            alert = alert_generator.review_alert(alert, reviewer, status_value, notes)
        
        return Response(AlertSerializer(alert).data)

    @action(detail=True, methods=['post'], url_path='assign')
    def assign_alert(self, request, alert_id=None):
        """
        Issue #39: Assign (or unassign, with assigned_to='') an alert to an
        investigator for triage. Logs the change in the case-history thread.
        """
        alert = self.get_object()
        serializer = AssignAlertSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        alert_generator = get_alert_generator()
        assigned_by = request.user.username if hasattr(request.user, 'username') else 'system'

        alert = alert_generator.assign_alert(
            alert,
            assigned_to=serializer.validated_data['assigned_to'],
            assigned_by=assigned_by,
            notes=serializer.validated_data.get('notes', ''),
        )

        return Response(AlertSerializer(alert).data)

    @action(detail=True, methods=['get', 'post'], url_path='comments')
    def comments(self, request, alert_id=None):
        """
        Issue #39: Case-history comment thread for an alert.
        GET  → list all comments (COMMENT / ASSIGNMENT / STATUS_CHANGE).
        POST → add a free-form investigation note ({"comment": "..."}).
        """
        alert = self.get_object()

        if request.method == 'GET':
            comments = alert.comments.all()
            return Response(AlertCommentSerializer(comments, many=True).data)

        comment_text = request.data.get('comment', '').strip()
        if not comment_text:
            return Response({'error': 'comment is required'}, status=status.HTTP_400_BAD_REQUEST)

        author = request.user.username if hasattr(request.user, 'username') else 'system'
        alert_generator = get_alert_generator()
        comment = alert_generator.add_comment(alert, author=author, comment=comment_text)

        return Response(AlertCommentSerializer(comment).data, status=status.HTTP_201_CREATED)

    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get alert statistics"""
        alert_generator = get_alert_generator()
        days = int(request.query_params.get('days', 30))
        stats = alert_generator.get_alerts_statistics(days=days)
        return Response(stats)
    
    @action(detail=False, methods=['get'])
    def open_count(self, request):
        """Get count of open alerts by severity"""
        alert_generator = get_alert_generator()
        counts = alert_generator.get_open_alerts_count()
        return Response(counts)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """
        Issue #15: Bulk export alerts to CSV or Excel (xlsx).

        Query params:
            export_format=csv  (default) — returns text/csv
            export_format=xlsx           — returns Excel workbook
            status=OPEN|...              — optional filter
            severity=HIGH|...            — optional filter

        Note: intentionally NOT named "format" — that query param is reserved
        by DRF for content negotiation, which runs before this view's code
        executes. Since only JSONRenderer is registered, ?format=csv/xlsx
        would make DRF raise Http404 before reaching this method.
        """
        export_format = request.query_params.get('export_format', 'csv').lower()
        queryset = self.get_queryset()

        # Optional filters
        alert_status = request.query_params.get('status')
        severity = request.query_params.get('severity')
        if alert_status:
            queryset = queryset.filter(status=alert_status)
        if severity:
            queryset = queryset.filter(severity=severity)

        queryset = queryset.select_related('customer', 'transaction').order_by('-created_at')

        headers = [
            'Alert ID', 'Severity', 'Status', 'Risk Score',
            'Customer ID', 'Customer Name',
            'Transaction ID', 'Amount', 'Currency',
            'Title', 'Description', 'Reviewed By', 'Created At',
        ]

        def row_for(alert):
            return [
                alert.alert_id,
                alert.severity,
                alert.status,
                str(alert.risk_score),
                alert.customer.customer_id,
                f"{alert.customer.first_name} {alert.customer.last_name}",
                alert.transaction.transaction_id,
                str(alert.transaction.amount),
                alert.transaction.currency,
                alert.title,
                alert.description,
                alert.reviewed_by or '',
                alert.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ]

        if export_format == 'xlsx':
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                return Response({'error': 'openpyxl not installed'}, status=500)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Alerts'

            # Header row with styling
            ws.append(headers)
            header_fill = PatternFill('solid', fgColor='1e3a5f')
            header_font = Font(bold=True, color='FFFFFF')
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for alert in queryset:
                ws.append(row_for(alert))

            # Auto-width columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="alerts_export.xlsx"'
            return response

        else:  # CSV
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="alerts_export.csv"'
            response.write('\ufeff')  # BOM for Excel UTF-8
            writer = csv.writer(response)
            writer.writerow(headers)
            for alert in queryset:
                writer.writerow(row_for(alert))
            return response


class DeviceViewSet(viewsets.ModelViewSet):
    """ViewSet for Device model (fraud/abuse device fingerprinting)"""
    queryset = Device.objects.all()
    serializer_class = DeviceSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'device_id'
    filterset_fields = ['device_type', 'is_emulator', 'is_rooted']
    search_fields = ['device_id', 'fingerprint_hash', 'ip_address']
    ordering_fields = ['first_seen_at', 'last_seen_at']
    ordering = ['-last_seen_at']

    @action(detail=True, methods=['get'])
    def customers(self, request, device_id=None):
        """Distinct customers that have transacted with this device (ring detection)."""
        device = self.get_object()
        customer_ids = Transaction.objects.filter(device=device).values_list(
            'customer_id', flat=True
        ).distinct()
        customers = Customer.objects.filter(id__in=customer_ids)
        return Response(CustomerSerializer(customers, many=True).data)


class MerchantViewSet(viewsets.ModelViewSet):
    """ViewSet for Merchant model (fraud/abuse merchant risk)"""
    queryset = Merchant.objects.all()
    serializer_class = MerchantSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'merchant_id'
    filterset_fields = ['category', 'is_active']
    search_fields = ['merchant_id', 'name']
    ordering_fields = ['risk_score', 'chargeback_rate', 'refund_rate', 'created_at']
    ordering = ['-risk_score']


class RiskScoreViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for RiskScore model (read-only)"""
    queryset = RiskScore.objects.all()
    serializer_class = RiskScoreSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['score_type']
    ordering_fields = ['calculated_at', 'score']
    ordering = ['-calculated_at']

    def get_queryset(self):
        queryset = RiskScore.objects.all()
        
        # Filter by customer
        customer_id = self.request.query_params.get('customer_id', None)
        if customer_id:
            queryset = queryset.filter(customer__customer_id=customer_id)
        
        # Filter by score type
        score_type = self.request.query_params.get('score_type', None)
        if score_type:
            queryset = queryset.filter(score_type=score_type)
        
        return queryset


class RuleViewSet(viewsets.ModelViewSet):
    """ViewSet for Rule model"""
    queryset = Rule.objects.all()
    serializer_class = RuleSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['status', 'rule_type']
    search_fields = ['name', 'description']
    ordering_fields = ['priority', 'created_at', 'risk_weight']
    ordering = ['priority', '-created_at']

    def get_queryset(self):
        queryset = Rule.objects.all()
        
        # Filter by status
        rule_status = self.request.query_params.get('status', None)
        if rule_status:
            queryset = queryset.filter(status=rule_status)
        
        # Filter by rule type
        rule_type = self.request.query_params.get('rule_type', None)
        if rule_type:
            queryset = queryset.filter(rule_type=rule_type)
        
        return queryset
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user.username if hasattr(self.request.user, 'username') else 'system')


class ReportViewSet(viewsets.ModelViewSet):
    """ViewSet for Report model"""
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'report_id'
    filterset_fields = ['report_type', 'status', 'file_format']
    search_fields = ['report_id', 'title', 'regulatory_body']
    ordering_fields = ['created_at', 'submitted_at']
    ordering = ['-created_at']

    def get_queryset(self):
        queryset = Report.objects.all()
        
        # Filter by report type
        report_type = self.request.query_params.get('report_type', None)
        if report_type:
            queryset = queryset.filter(report_type=report_type)
        
        # Filter by status
        report_status = self.request.query_params.get('status', None)
        if report_status:
            queryset = queryset.filter(status=report_status)
        
        return queryset
    
    @action(detail=False, methods=['post'], url_path='generate')
    def generate_report(self, request):
        """Generate a new report"""
        serializer = GenerateReportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        report_type = serializer.validated_data['report_type']
        period_start = serializer.validated_data['period_start']
        period_end = serializer.validated_data['period_end']
        file_format = serializer.validated_data.get('format', 'JSON')
        submitted_by = request.user.username if hasattr(request.user, 'username') else 'system'
        
        report_generator = get_report_generator()
        
        try:
            if report_type == 'SAR':
                # Get alerts in the period
                alerts = Alert.objects.filter(
                    created_at__gte=period_start,
                    created_at__lte=period_end
                )
                
                if not alerts.exists():
                    return Response(
                        {'error': 'No alerts found in the specified period'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                report = report_generator.generate_sar(
                    alerts=list(alerts),
                    period_start=period_start,
                    period_end=period_end,
                    submitted_by=submitted_by
                )
            
            elif report_type == 'CTR':
                threshold = serializer.validated_data.get('threshold', None)
                if not threshold:
                    return Response(
                        {'error': 'Threshold is required for CTR reports'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Get transactions exceeding threshold
                transactions = Transaction.objects.filter(
                    transaction_date__gte=period_start,
                    transaction_date__lte=period_end,
                    amount__gte=threshold,
                    status='COMPLETED'
                )
                
                if not transactions.exists():
                    return Response(
                        {'error': 'No transactions found exceeding the threshold'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                report = report_generator.generate_ctr(
                    transactions=list(transactions),
                    period_start=period_start,
                    period_end=period_end,
                    threshold=threshold,
                    submitted_by=submitted_by
                )
            
            else:
                return Response(
                    {'error': f'Report type {report_type} not implemented'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Export to requested format
            if file_format == 'JSON':
                report_generator.export_report_json(report)
            elif file_format == 'CSV':
                report_generator.export_report_csv(report)
            elif file_format == 'PDF':
                report_generator.export_report_pdf(report)
            
            return Response(ReportSerializer(report).data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"Error generating report: {str(e)}")
            return Response(
                {'error': f'Error generating report: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def download(self, request, report_id=None):
        """Download report file"""
        report = self.get_object()
        
        if not report.file_path:
            return Response(
                {'error': 'Report file not generated'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        try:
            return FileResponse(
                open(report.file_path, 'rb'),
                as_attachment=True,
                filename=f"{report.report_id}.{report.file_format.lower()}"
            )
        except FileNotFoundError:
            return Response(
                {'error': 'Report file not found'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['post'])
    def submit(self, request, report_id=None):
        """Submit report to regulatory body"""
        report = self.get_object()
        regulatory_body = request.data.get('regulatory_body', '')
        
        report_generator = get_report_generator()
        report = report_generator.submit_report(report, regulatory_body)
        
        return Response(ReportSerializer(report).data)


# --- Audit log (read-only, paginated, for compliance) ---


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """Read-only list of audit log entries (compliance API)."""
    queryset = AuditLog.objects.all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['method', 'path', 'user', 'status_code']
    ordering_fields = ['timestamp', 'path', 'status_code']
    ordering = ['-timestamp']


# --- Health & readiness (no auth for load balancers) ---

from rest_framework.views import APIView
from django.db import connection


class HealthView(APIView):
    """GET /api/health/ — liveness (app is up)."""
    permission_classes = [AllowAny]
    throttle_classes = []  # No rate limit for load balancers

    def get(self, request):
        return Response({'status': 'ok', 'service': 'didebaan'})


class ReadyView(APIView):
    """GET /api/ready/ — readiness (app can serve traffic, DB reachable)."""
    permission_classes = [AllowAny]
    throttle_classes = []  # No rate limit for load balancers

    def get(self, request):
        try:
            with connection.cursor() as cursor:
                cursor.execute('SELECT 1')
            return Response({'status': 'ready', 'database': 'ok'})
        except Exception as e:
            return Response(
                {'status': 'not_ready', 'database': 'error', 'detail': str(e)},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )

